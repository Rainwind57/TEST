"""回测/选股报告服务：HTML（含图表）/ Excel / PDF 渲染 + 报告历史落盘。

供 routers/reports.py（导出接口）、selection.py / ml.py（回测完成自动存档）复用。
"""
import base64
import io
import os
import datetime
import logging
import uuid

from . import db

logger = logging.getLogger(__name__)

_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
_ECHARTS_PATH = os.path.join(_STATIC_DIR, "echarts.min.js")


def _echarts_script_tag() -> str:
    """返回 ECharts 加载脚本标签。

    优先从 backend/app/static/echarts.min.js 内联 base64（完全离线、自包含），
    若文件不存在则降级为多 CDN 链（国内→国际）。
    """
    if os.path.exists(_ECHARTS_PATH):
        with open(_ECHARTS_PATH, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        return f'<script src="data:text/javascript;base64,{b64}"></script>'

    # 降级：多 CDN 容错链（cdn.bootcdn.net 在国内较稳，再备 unpkg / cdnjs）
    return (
        '<script>'
        '(function(){'
        'var u=["https://cdn.bootcdn.net/ajax/libs/echarts/5.5.0/echarts.min.js",'
        '"https://unpkg.com/echarts@5.5.0/dist/echarts.min.js",'
        '"https://cdnjs.cloudflare.com/ajax/libs/echarts/5.5.0/echarts.min.js"],'
        'i=0;function t(){if(i>=u.length){'
        'document.getElementById("chart-fallback").style.display="block";return}'
        'var s=document.createElement("script");s.src=u[i++];'
        's.onerror=t;document.head.appendChild(s)}t()'
        '})()'
        '</script>'
    )

REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
os.makedirs(REPORT_DIR, exist_ok=True)


def _htmlescape(s: str) -> str:
    """HTML 实体转义：防止用户输入注入脚本。"""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#x27;")

def _fmt(v, pct=False):
    if v is None:
        return "-"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    return f"{v*100:.2f}%" if pct else f"{v:.4f}"


def payload_from_result(result: dict) -> dict:
    """从回测结果 dict 提取报告载荷（前端导出与后端存档共用同一结构）。"""
    cfg = result.get("config") or {}
    return {
        "factorLabel": result.get("factorLabel", "回测"),
        "config": cfg,
        "metrics": result.get("metrics") or {},
        "benchmark": result.get("benchmark"),
        "groupSummary": result.get("groupSummary") or [],
        "longShort": result.get("longShort") or [],
        "positionLedger": result.get("positionLedger") or [],
        "icSeries": result.get("icSeries") or [],
        "survivorshipBiasWarning": result.get("survivorshipBiasWarning", ""),
        "snapshotWarning": result.get("snapshotWarning", ""),
        "snapshotStartNote": result.get("snapshotStartNote", ""),
        "inSampleWarning": result.get("inSampleWarning", ""),
        "histWarning": result.get("histWarning", ""),
        "actualHistDays": result.get("actualHistDays"),
        "effectiveStart": result.get("effectiveStart"),
        "effectiveEnd": result.get("effectiveEnd"),
        "longOnly": result.get("longOnly", True),
        "longOnlyNote": result.get("longOnlyNote", ""),
        "longShortNote": result.get("longShortNote", ""),
        "featureImportance": result.get("featureImportance") or [],
        "signalMode": result.get("signalMode", "group"),
        "bullRule": result.get("bullRule", ""),
        "bearRule": result.get("bearRule", ""),
        "topAttribution": result.get("topAttribution"),
        # 方向选择 + 报告补全（分层回测优化）
        "direction": result.get("direction", cfg.get("direction", "long_short")),
        "icIr": result.get("icIr"),
        "icTStat": result.get("icTStat"),
        "icPValue": result.get("icPValue"),
        "yearlyReturns": result.get("yearlyReturns") or [],
        "stockContribution": result.get("stockContribution") or [],
        "bucketDates": result.get("bucketDates") or [],
        # 空结果结构化信息（U2：回测无有效样本时明确提示，而非静默空白）
        "ok": result.get("ok", True),
        "reason": result.get("reason"),
        "error": result.get("error"),
        "hint": result.get("hint"),
        "requested": result.get("requested"),
        "actualWindow": result.get("actualWindow"),
        # 模型详情：config 中携带的模型 meta（routers/ml.py 回测时注入）
        "_modelMeta": cfg.get("_modelMeta") if cfg.get("modelId") else None,
    }


def render_html(payload: dict, title: str = "回测报告") -> str:
    """自包含 HTML：KPI 卡片 + 表格 + ECharts（CDN）净值/回撤/分组/IC 图。"""
    import json as _json
    m = payload.get("metrics") or {}
    bench = payload.get("benchmark")
    cfg = payload.get("config") or {}
    direction = payload.get("direction") or cfg.get("direction") or "long_short"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # 回测概要信息块（起止日 / 历史长度 / 成本 / 候选池 → 解决"报告看不出回测区间"的历史问题）
    start = cfg.get("startDate") or "-"
    end = cfg.get("endDate") or "-"
    hist_val = cfg.get("hist") or "-"
    board_val = cfg.get("board") or "-"
    pool_size = cfg.get("poolSize") or "-"
    n_val = cfg.get("n") or "-"
    cost = f"佣金{_fmt(cfg.get('commissionRate') or 0.00025)} "
    cost += f"印花{_fmt(cfg.get('stampDuty') or 0.001)} "
    cost += f"滑点{_fmt(cfg.get('slippage') or 0.001)}"
    strategy_kind = "ML模型" if cfg.get("modelId") else "技术因子"
    strategy_name = cfg.get("modelId") or payload.get("factorLabel") or "-"
    if strategy_kind == "ML模型":
        strategy_name = cfg.get("modelId") or "-"

    # 警告横幅：前视偏差 / 生存偏差 / 多空不可实盘 / 历史长度钳制
    warnings = []
    sw = payload.get("snapshotWarning", "")
    if sw:
        warnings.append(f'<div class="warn-banner" style="border-color:#e6a817;background:#1a1410">⚠️ 前视偏差警告：{_htmlescape(sw)}</div>')
    sn = payload.get("snapshotStartNote", "")
    if sn:
        warnings.append(f'<div class="warn-banner" style="border-color:#e6a817;background:#1a1410">⚠️ 回测起点调整：{_htmlescape(sn)}</div>')
    isw = payload.get("inSampleWarning", "")
    if isw:
        warnings.append(f'<div class="warn-banner" style="border-color:#4f8cff;background:#101a2a">ℹ️ 样本内回测：{_htmlescape(isw)}</div>')
    sb = payload.get("survivorshipBiasWarning", "")
    if sb:
        warnings.append(f'<div class="warn-banner">⚠️ 生存偏差：{_htmlescape(sb)}</div>')
    hw = payload.get("histWarning", "")
    if hw:
        warnings.append(f'<div class="info-banner">📌 历史长度调整：{_htmlescape(hw)}</div>')
    lo = payload.get("longOnly", True)
    if lo:
        note = payload.get("longOnlyNote", "")
        if note:
            warnings.append(f'<div class="info-banner">📌 {_htmlescape(note)}</div>')
    else:
        note = payload.get("longShortNote", "")
        if note:
            warnings.append(f'<div class="warn-banner" style="border-color:#e6a817;background:#1a1410">⚠️ {_htmlescape(note)}</div>')
    # 空结果提示：回测无有效样本时明确告知，避免静默空白报告
    ls = payload.get("longShort") or []
    if payload.get("ok") is False:
        reason = payload.get("reason") or "NO_VALID_REBALANCE"
        hint = payload.get("hint") or "回测区间无有效调仓日，请增大历史长度(hist)或后移终止日。"
        req = payload.get("requested") or {}
        aw = payload.get("actualWindow") or {}
        req_line = ""
        if req.get("start") or req.get("end") or req.get("hist"):
            req_line = f"请求区间 {req.get('start') or '-'} ~ {req.get('end') or '-'}，历史 {req.get('hist') or '-'} 日"
        aw_line = ""
        if aw.get("start") and aw.get("end"):
            aw_line = f"实际K线覆盖 {aw.get('start')} ~ {aw.get('end')}"
        warnings.append(
            f'<div class="warn-banner">⚠️ 回测无有效样本（{_htmlescape(reason)}）：{_htmlescape(hint)}'
            + (f'<br>{_htmlescape(req_line)}；{_htmlescape(aw_line)}' if (req_line or aw_line) else '')
            + '</div>'
        )
    elif not ls:
        eff_start = payload.get("effectiveStart") or ""
        eff_end = payload.get("effectiveEnd") or ""
        actual_hist = payload.get("actualHistDays") or "-"
        detail = ""
        if eff_start and eff_end:
            detail = f"，有效区间 {eff_start} ~ {eff_end}"
        if actual_hist and actual_hist != "-":
            detail += f"，实际拉取 {actual_hist} 日K线"
        warnings.append(
            f'<div class="warn-banner">⚠️ 回测无有效样本：所选时间段与历史长度不匹配，'
            f'调仓日截面均未凑够最低样本数{detail}。'
            f'请检查回测起止日期是否落在K线覆盖范围内，或增大历史长度(hist)参数。</div>'
        )

    model_meta = payload.get("_modelMeta")

    # 因子说明块（技术因子回测无模型详情时，补充因子定义/方向/逻辑）
    factor_html = ""
    if not model_meta and payload.get("factorLabel"):
        factor_label = payload.get("factorLabel", "")
        factor_display_label = factor_label
        try:
            from .factors import FACTORS, SNAPSHOT_FACTORS
            # 优先用 config 中的英文因子 key 查表，退化为按中文 label 遍历匹配
            factor_key = cfg.get("factor", "")
            fdef = FACTORS.get(factor_key) or SNAPSHOT_FACTORS.get(factor_key)
            if not fdef:
                all_factors = {**FACTORS, **SNAPSHOT_FACTORS}
                fdef = next((v for v in all_factors.values() if v.get("label") == factor_label), None)
        except Exception:
            fdef = None
        if fdef:
            factor_html = (
                f'<div class="card"><h3>因子说明</h3><table class="summary-table">'
                f'<tr><td>因子名称</td><td>{fdef.get("label", factor_display_label)}</td>'
                f'<td>方向</td><td>{fdef.get("direction", "-")}</td></tr>'
                f'<tr><td>分组</td><td>{fdef.get("group", "-")}</td>'
                f'<td>窗口</td><td>{fdef.get("window", "-")}</td></tr>'
                f'</table></div>'
            )

    # 模型详情块
    model_html = ""
    if model_meta:
        model_type = model_meta.get("modelType", "gbdt")
        n_features = len(model_meta.get("featureNames") or [])
        bp = model_meta.get("bestParams")
        hp_str = ""
        if bp:
            parts = []
            if "n_estimators" in bp:
                parts.append(f"n_estimators={bp['n_estimators']}")
            if "max_depth" in bp:
                parts.append(f"max_depth={bp['max_depth']}")
            if "learning_rate" in bp:
                parts.append(f"learning_rate={bp['learning_rate']}")
            hp_str = " · ".join(parts)
        top_features = (model_meta.get("featureImportance") or [])[:5]
        ft_rows = "".join(
            f"<tr><td>{f['feature']}</td><td>{f['importance']:.4f}</td></tr>"
            for f in top_features
        )
        model_html = (
            f'<div class="card"><h3>模型详情</h3><table class="summary-table">'
            f'<tr><td>模型类型</td><td>{model_type}</td>'
            f'<td>特征数</td><td>{n_features}</td></tr>'
            + (f'<tr><td>超参数</td><td colspan="3">{hp_str}</td></tr>' if hp_str else "") +
            f'</table>'
            f'<h4 style="margin-top:12px">Top 特征重要性</h4>'
            f'<table><tr><th>特征</th><th>重要性</th></tr>{ft_rows}</table></div>'
        )

    # 调仓方式文字
    groups_val = cfg.get("groups") or "-"
    signal_mode = payload.get("signalMode", "group")
    bull_rule = payload.get("bullRule", "")
    bear_rule = payload.get("bearRule", "")
    cost_note = ""
    if cfg.get("applyCost"):
        cost_note = (
            "交易成本（佣金+印花税+滑点）已计入。高频调仓（n≤3日）时成本侵蚀显著，"
            "建议关注减成本后指标（如净Sharpe与毛Sharpe的差异）。"
        )
    else:
        cost_note = "⚠️ 未计入交易成本，实际收益需扣除佣金/印花税/滑点。"
    if signal_mode == "rule":
        rebalance_desc = (
            f"离散规则信号 · 每 {n_val} 交易日按规则判定买卖"
            + (f" · 看多「{_htmlescape(bull_rule)}」" if bull_rule else "")
            + (f" · 看空「{_htmlescape(bear_rule)}」" if bear_rule else "")
            + (f" · T+1开盘入场 · {cost_note}")
        )
        warnings.append(
            f'<div class="info-banner">📌 信号模式：离散规则引擎（非分位分组）。'
            f'看多规则「{_htmlescape(bull_rule or "-")}」/ 看空规则「{_htmlescape(bear_rule or "-")}」'
            f'驱动逐期买卖信号与物理持仓。</div>'
        )
    else:
        if direction == "long_only":
            dir_desc = "做多Top组（最高分，仅多头可实盘）"
        elif direction == "short_only":
            dir_desc = "做空Bottom组（最低分，仅空头需融券）"
        else:
            dir_desc = "做多Top组 / 做空Bottom组（研究用多空组合）"
        rebalance_desc = (
            f"分层回测 · 每 {n_val} 交易日按因子值全市场排序分 {groups_val} 组 · "
            + dir_desc + (f" · T+1开盘入场 · {cost_note}")
        )

    warnings_html = "\n".join(warnings) if warnings else ""

    # M8 预测归因：最后一个调仓日 Top 看多/看空个股各特征截面分位
    attribution_html = ""
    attr = payload.get("topAttribution")
    if attr:
        f_labels = attr.get("featureLabels") or attr.get("featureNames") or []
        f_keys = attr.get("featureNames") or []

        def _attr_rows(items, side_label):
            rows = ""
            for it in items:
                pcts = it.get("featurePcts") or {}
                cells = "".join(
                    f'<td>{pcts[k]:.2f}</td>' if pcts.get(k) is not None else '<td>-</td>'
                    for k in f_keys
                )
                rows += (f'<tr><td>{side_label}</td><td>{_htmlescape(it.get("code", ""))}</td>'
                         f'<td>{_fmt(it.get("score"))}</td>{cells}</tr>')
            return rows

        head = "".join(f'<th>{_htmlescape(str(lbl))}</th>' for lbl in f_labels)
        attr_date = attr.get("date") or ""
        attribution_html = (
            f'<div class="card"><h3>预测归因（{_htmlescape(attr_date)}）</h3>'
            f'<p class="sub" style="margin:4px 0 8px">最后一个调仓日 Top 看多/看空个股的各特征截面分位'
            f'（0~1，越接近 1 代表该特征在当日截面越靠前），辅助理解信号来源。</p>'
            f'<table><tr><th>方向</th><th>代码</th><th>预测分</th>{head}</tr>'
            f'{_attr_rows(attr.get("longs") or [], "看多")}'
            f'{_attr_rows(attr.get("shorts") or [], "看空")}'
            f'</table></div>'
        )

    # 有效区间（从回测结果中读取，解决"报告看不出回测区间"的历史问题）
    eff_start = payload.get("effectiveStart") or cfg.get("effectiveStart") or "-"
    eff_end = payload.get("effectiveEnd") or cfg.get("effectiveEnd") or "-"
    actual_hist = payload.get("actualHistDays") or cfg.get("actualHistDays") or hist_val

    summary = (
        f'{warnings_html}'
        f'{factor_html}'
        f'{model_html}'
        f'<div class="card"><h3>回测概要</h3><table class="summary-table">'
        f'<tr><td>调仓方式</td><td colspan="3">{rebalance_desc}</td></tr>'
    f'<tr><td>策略来源</td><td>{_htmlescape(strategy_kind)}</td>'
    f'<td>策略名/模型ID</td><td>{_htmlescape(strategy_name)}</td></tr>'
        f'<tr><td>回测区间</td><td>{start} ~ {end}</td>'
        f'<td>历史长度(hist)</td><td>{hist_val}</td></tr>'
        f'<tr><td>有效调仓区间</td><td>{eff_start} ~ {eff_end}</td>'
        f'<td>实际拉取K线天数</td><td>{actual_hist}</td></tr>'
        f'<tr><td>持有期(n)</td><td>{n_val}</td>'
        f'<td>候选池规模</td><td>{pool_size}</td></tr>'
        f'<tr><td>板块</td><td>{board_val}</td>'
        f'<td>成本</td><td>{cost}</td></tr>'
        f'</table></div>'
    )

    kpi = lambda label, val, pct=False: (
        f'<div class="kpi"><div class="n">{_fmt(val, pct)}</div>'
        f'<div class="l">{label}</div></div>'
    )
    if direction == "long_only":
        kpi_label = "多头"
    elif direction == "short_only":
        kpi_label = "空头"
    else:
        kpi_label = "多空"
    kpis = "".join([
        kpi(f"{kpi_label}累计收益", m.get("cumulativeReturn"), True),
        kpi(f"{kpi_label}年化", m.get("annualizedReturn"), True),
        kpi("年化波动", m.get("annualizedVolatility"), True),
        kpi("Sharpe", m.get("sharpe")),
        kpi("Sortino", m.get("sortino")),
        kpi("最大回撤", m.get("maxDrawdown"), True),
        kpi("Calmar", m.get("calmar")),
        kpi("胜率", m.get("winRate"), True),
    ])

    rows = ""
    for g in payload.get("groupSummary") or []:
        rows += (
            f"<tr><td>第{g.get('group')}组</td>"
            f"<td>{_fmt(g.get('avgReturn'), True)}</td>"
            f"<td>{_fmt(g.get('cumReturn'), True)}</td>"
            f"<td>{_fmt(g.get('annualizedReturn'), True)}</td>"
            f"<td>{_fmt(g.get('sharpe'))}</td>"
            f"<td>{_fmt(g.get('maxDrawdown'), True)}</td>"
            f"<td>{g.get('sample')}</td></tr>"
        )

    bench_html = ""
    if bench:
        bench_html = (
            '<div class="card"><h3>基准对比</h3><table>'
            "<tr><th>指标</th><th>策略</th><th>基准</th></tr>"
            f"<tr><td>累计收益</td><td>{_fmt(m.get('cumulativeReturn'), True)}</td>"
            f"<td>{_fmt(bench.get('cumulativeReturn'), True)}</td></tr>"
            f"<tr><td>年化收益</td><td>{_fmt(m.get('annualizedReturn'), True)}</td>"
            f"<td>{_fmt(bench.get('annualizedReturn'), True)}</td></tr>"
            f"<tr><td>Sharpe</td><td>{_fmt(m.get('sharpe'))}</td>"
            f"<td>{_fmt(bench.get('sharpe'))}</td></tr>"
            f"<tr><td>策略 Alpha</td><td>{_fmt(bench.get('strategyAlpha', bench.get('alpha')))}</td><td>-</td></tr>"
            f"<tr><td>策略 Beta</td><td>{_fmt(bench.get('strategyBeta', bench.get('beta')))}</td><td>-</td></tr>"
            "</table></div>"
        )

    # IC 统计
    ic_stats = ""
    ics: list[float] = []
    mean_ic = 0.0
    pos = 0.0
    if payload.get("icSeries"):
        ics = [p.get("ic") for p in payload["icSeries"] if p.get("ic") is not None]
        rankics = [p.get("rankIc") for p in payload["icSeries"] if p.get("rankIc") is not None]
        if ics:
            mean_ic = sum(ics) / len(ics)
            mean_rank_ic = sum(rankics) / len(rankics) if rankics else 0.0
            pos = sum(1 for v in ics if v > 0) / len(ics)
            icir = payload.get("icIr")
            tstat = payload.get("icTStat")
            pval = payload.get("icPValue")
            sig_str = ""
            if tstat is not None:
                sig_str = f" · t值 {_fmt(tstat)}"
            if pval is not None:
                sig_str += f" · p值 {_fmt(pval)}（{'显著' if pval < 0.05 else '不显著'}）"
            ic_stats = (f'<div class="card"><h3>IC 统计</h3><p>'
                        f'平均 IC：{_fmt(mean_ic)} · RankIC：{_fmt(mean_rank_ic)} · IC 胜率：{pos*100:.1f}% · 截面数：{len(ics)}'
                        + (f' · ICIR：{_fmt(icir)}' if icir is not None else '')
                        + sig_str + '</p></div>')

    # 结果解读
    sharpe = m.get("sharpe") or 0
    mdd = m.get("maxDrawdown") or 0
    wr = m.get("winRate") or 0
    ic_val = mean_ic
    ic_pos = pos

    sharpe_grade = "优秀（≥1.5）" if sharpe >= 1.5 else "良好（1.0~1.5）" if sharpe >= 1.0 else "一般（0.5~1.0）" if sharpe >= 0.5 else "偏低（<0.5）"
    mdd_grade = "较大（<-30%）" if mdd < -0.3 else "中等（-20%~-30%）" if mdd < -0.2 else "可控（>-20%）"
    ic_grade = "有效（均值>0.03）" if ic_val > 0.03 else "中等（0.01~0.03）" if ic_val > 0.01 else "偏弱（<0.01）"
    conclusion = []
    if sharpe >= 1.0 and mdd > -0.25:
        conclusion.append("策略整体风险调整后收益表现良好，回撤可控，适合作为组合基础仓位。")
    elif sharpe >= 0.5:
        conclusion.append("策略有一定选股能力但回撤偏高，建议结合仓位管理或止损机制使用。")
    else:
        conclusion.append("策略信号偏弱，建议优化因子组合或扩大候选池后再评估。")
    if ic_val > 0.02:
        conclusion.append("IC 均值表明因子对截面收益有区分度，可配合其他因子构建多因子模型。")
    if len(ics) < 20 and len(ics) > 0:
        conclusion.append(f"调仓次数仅 {len(ics)} 次，样本偏少，绩效指标统计显著性不足。")

    interp_html = (
        f'<div class="card"><h3>结果解读</h3>'
        f'<p style="line-height:1.8;color:#cfe0ff">Sharpe {sharpe_grade} · '
        f'最大回撤 {mdd_grade} · 胜率 {wr*100:.1f}% · '
        f'因子IC {ic_grade}。{" ".join(conclusion)}</p>'
        f'</div>'
    )

    # B4：交易成本影响卡片
    if m.get("applyCost"):
        cost_card = (
            f'<div class="card"><h3>交易成本影响</h3><table class="summary-table">'
            f'<tr><td>单边往返成本率</td><td>{_fmt(m.get("costRate"), True)}</td>'
            f'<td>年化换手</td><td>{_fmt(m.get("annualizedTurnover"))} 股/年</td></tr>'
            f'<tr><td>毛Sharpe（除成本）</td><td>{_fmt(m.get("grossSharpe"))}</td>'
            f'<td>净Sharpe（含成本）</td><td>{_fmt(m.get("sharpe"))}</td></tr>'
            f'<tr><td>成本侵蚀Sharpe</td><td colspan="3">{_fmt(m.get("costErosion"))}</td></tr>'
            f'</table></div>'
        )
    else:
        cost_card = ('<div class="card"><h3>交易成本影响</h3>'
                     '<p class="sub">未计入交易成本（applyCost=False），Sharpe 等指标为毛收益口径。</p></div>')

    # B7：分年度绩效
    yrows = ""
    for y in payload.get("yearlyReturns") or []:
        yrows += (f"<tr><td>{y.get('year')}</td>"
                  f"<td>{_fmt(y.get('return'), True)}</td>"
                  f"<td>{y.get('periods')}</td></tr>")
    yearly_html = ""
    if yrows:
        yearly_html = ('<div class="card"><h3>分年度绩效</h3><table>'
                       '<tr><th>年份</th><th>年度收益</th><th>调仓次数</th></tr>'
                       + yrows + '</table></div>')

    # B5：个股贡献 Top30
    crows = ""
    for sc in payload.get("stockContribution") or []:
        crows += (f"<tr><td>{_htmlescape(str(sc.get('code', '')))}</td>"
                  f"<td>{_htmlescape(str(sc.get('name', '')))}</td>"
                  f"<td>{_fmt(sc.get('contribution'), True)}</td></tr>")
    contrib_html = ""
    if crows:
        contrib_html = ('<div class="card"><h3>个股贡献 Top30</h3>'
                        '<p class="sub" style="margin:4px 0 8px">按绝对累计贡献排序（多头 +收益 / 空头 -收益）。</p>'
                        '<table><tr><th>代码</th><th>名称</th><th>累计贡献</th></tr>'
                        + crows + '</table></div>')

    # B5：持仓明细（最近 10 期）
    holdings_html = ""
    pl = payload.get("positionLedger") or []
    if pl:
        hrows = ""
        for p in list(reversed(pl))[:10]:
            lng = ", ".join(p.get("long") or []) or "-"
            sht = ", ".join(p.get("short") or []) or "-"
            hrows += (f"<tr><td>{_htmlescape(str(p.get('date', '')))}</td>"
                      f"<td>{_htmlescape(lng)}</td>"
                      f"<td>{_htmlescape(sht)}</td></tr>")
        holdings_html = ('<div class="card"><h3>持仓明细（最近 10 期）</h3>'
                         '<p class="sub" style="margin:4px 0 8px">每期多/空持仓代码，最新在上。</p>'
                         '<table><tr><th>调仓日</th><th>多头持仓</th><th>空头持仓</th></tr>'
                         + hrows + '</table></div>')

    # B11：多空收益分解（多头腿 vs 空头腿累计贡献，帮助理解对冲来源）
    decomp_html = ""
    if direction == "long_short":
        ls_points = payload.get("longShort") or []
        if ls_points:
            last_p = ls_points[-1]
            top_cum = last_p.get("topCum")
            bottom_cum = last_p.get("bottomCum")
            if top_cum is not None and bottom_cum is not None:
                net = last_p.get("cum")
                decomp_html = (
                    '<div class="card"><h3>多空收益分解</h3><table class="summary-table">'
                    f'<tr><td>多头腿累计收益</td><td>{_fmt(top_cum, True)}</td>'
                    f'<td>空头腿累计收益</td><td>{_fmt(bottom_cum, True)}</td></tr>'
                    f'<tr><td>多空合并累计</td><td colspan="3">{_fmt(net, True)}'
                    f'（= 多头腿 − 空头腿，空头腿为做空收益）</td></tr>'
                    '</table></div>'
                )

    if direction == "long_only":
        ls_title = "多头累计收益与回撤"
        ls_desc = "蓝线为多头组合累计收益率（%），橙色阴影为回撤。长期右上→策略有效。"
        main_series_name = "多头累计"
    elif direction == "short_only":
        ls_title = "空头累计收益与回撤"
        ls_desc = "蓝线为空头组合累计收益率（%），橙色阴影为回撤。长期右上→策略有效。"
        main_series_name = "空头累计"
    else:
        ls_title = "多空累计收益与回撤"
        ls_desc = "蓝线为多空组合累计收益率（%），绿线多头腿、紫线空头腿，橙色阴影为回撤。长期右上→策略有效。"
        main_series_name = "多空累计"

    ls_json = _json.dumps(payload.get("longShort") or [], ensure_ascii=False)
    gs_json = _json.dumps(payload.get("groupSummary") or [], ensure_ascii=False)
    ic_json = _json.dumps(payload.get("icSeries") or [], ensure_ascii=False)
    fi_json = _json.dumps(payload.get("featureImportance") or [], ensure_ascii=False)
    bd_json = _json.dumps(payload.get("bucketDates") or [], ensure_ascii=False)
    dir_json = _json.dumps(direction)
    label = (payload.get("factorLabel") or title).replace("<", "&lt;").replace(">", "&gt;")

    return f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8">
<title>{label} · {title}</title>
{_echarts_script_tag()}
<style>
body{{font-family:-apple-system,"Segoe UI","PingFang SC",sans-serif;background:#0b1020;color:#e6ebf5;margin:0;padding:32px}}
h1{{font-size:24px;border-left:4px solid #4f8cff;padding-left:10px}}
h3{{color:#cfe0ff;margin-top:24px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin:16px 0}}
.kpi{{background:#121a30;border:1px solid #23304f;border-radius:10px;padding:14px}}
.kpi .n{{font-size:22px;font-weight:700;color:#4f8cff}}
.kpi .l{{color:#8e9bbd;font-size:12px;margin-top:4px}}
.card{{background:#121a30;border:1px solid #23304f;border-radius:10px;padding:16px;margin:14px 0}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th,td{{border:1px solid #23304f;padding:8px;text-align:left}}
th{{background:#16213c;color:#cfe0ff}}
.chart{{width:100%;height:340px}}
.sub{{color:#8e9bbd;font-size:13px}}
.summary-table td{{padding:6px 12px; vertical-align:top;}}
.summary-table td:first-child,.summary-table td:nth-child(3){{color:#8e9bbd; font-size:12px; width:100px;}}
.warn-banner{{background:#1a1410;border:1px solid #e6604d;border-radius:8px;padding:12px 16px;margin:12px 0;color:#ffa39e;font-size:13px;line-height:1.6}}
.info-banner{{background:#101a25;border:1px solid #4f8cff;border-radius:8px;padding:12px 16px;margin:12px 0;color:#a0c4ff;font-size:13px;line-height:1.6}}
@media print{{.chart{{height:300px}} body{{background:#fff;color:#000}} .kpi,.card{{border-color:#ccc;background:#fff}} th{{background:#f0f0f0;color:#000}} h1,h3{{color:#000}}}}
	</style></head><body>
		<div id="chart-fallback" style="display:none;background:#2a1a1a;border:1px solid #ff4d4f;border-radius:8px;padding:16px;margin:16px 0;color:#ffa39e;font-size:13px">⚠️ 图表加载失败，请检查网络连接后刷新页面，或导出 Excel/PDF 格式查看数据</div>
		<h1>{label}</h1>
	<div class="sub">生成时间 {now} · 调仓次数 {m.get('rebalanceCount','-')} · 成本率 {_fmt(m.get('costRate'))}</div>
	{summary}
	<div class="grid">{kpis}</div>
{bench_html}
<div class="card"><h3>分组收益</h3><table>
<tr><th>分组</th><th>平均收益</th><th>累计收益</th><th>年化收益</th><th>Sharpe</th><th>最大回撤</th><th>样本数</th></tr>{rows}</table></div>
{ic_stats}
{interp_html}
{attribution_html}
{cost_card}
{yearly_html}
{contrib_html}
{holdings_html}
{decomp_html}
<div class="card"><h3>{ls_title}</h3><p class="sub" style="margin:4px 0 8px">{ls_desc}</p><div id="chartLS" class="chart"></div></div>
<div class="card"><h3>分组平均收益</h3><p class="sub" style="margin:4px 0 8px">柱状图为各分组的平均持仓期收益，分组1→N 单调→因子区分度好。</p><div id="chartGroup" class="chart"></div></div>
<div class="card"><h3>分组净值曲线</h3><p class="sub" style="margin:4px 0 8px">各组累计净值（%），组间分层越清晰→因子区分度越强。</p><div id="chartGroupEq" class="chart"></div></div>
<div class="card"><h3>IC / RankIC 时序列</h3><p class="sub" style="margin:4px 0 8px">IC（皮尔逊）/ RankIC（斯皮尔曼）时序。>0 天数多→因子预测方向稳，|IC|大→区分度强。</p><div id="chartIC" class="chart"></div></div>
<div class="card" id="fiCard" style="display:none"><h3>特征贡献</h3><p class="sub" style="margin:4px 0 8px">模型特征重要性 Top-N（条形图），横向越长→该因子对预测分贡献越大。</p><div id="chartFI" class="chart"></div></div>
<div class="card" id="chartPosCard" style="display:none"><h3>持仓与换手变化</h3><p class="sub" style="margin:4px 0 8px">每期多头/空头持仓数量及换手股数。换手高→模型对不同股票区分度强、信号变化活跃。</p><div id="chartPos" class="chart"></div></div>
<div class="card" id="rebalanceCard" style="display:none"><h3>调仓信号明细</h3><p class="sub" style="margin:4px 0 8px">每期相对上期的买入/卖出股票（最近 10 期）。多头买入=新增做多，多头卖出=平多；空头同理。</p><div id="rebalanceTable"></div></div>
<script>
const LS={ls_json}, GS={gs_json}, IC={ic_json}, BD={bd_json}, DIR={dir_json};
const dates=LS.map(p=>p.date);
const cum=LS.map(p=>Number((p.cum*100).toFixed(2)));
const top=LS.map(p=>Number(((p.topCum??p.cum)*100).toFixed(2)));
const bottom=LS.map(p=>Number(((p.bottomCum??0)*100).toFixed(2)));
let peak=-Infinity, dd=[];
for(const c of cum){{peak=Math.max(peak,c); dd.push(Number(((c-peak)/100*100).toFixed(2)));}}
const lsSeries=[{{name:'{main_series_name}',type:'line',data:cum,smooth:true,itemStyle:{{color:'#4f8cff'}},areaStyle:{{color:'rgba(79,140,255,.12)'}}}}];
if(DIR==='long_short'){{lsSeries.push({{name:'多头腿',type:'line',data:top,smooth:true,itemStyle:{{color:'#21c08b'}}}});lsSeries.push({{name:'空头腿',type:'line',data:bottom,smooth:true,itemStyle:{{color:'#6c5ce7'}}}});}}
lsSeries.push({{name:'回撤',type:'line',yAxisIndex:1,data:dd,itemStyle:{{color:'#f39c12'}}}});
echarts.init(document.getElementById('chartLS')).setOption({{
  tooltip:{{trigger:'axis'}}, legend:{{textStyle:{{color:'#8e9bbd'}},top:0}},
  grid:{{left:60,right:40,top:40,bottom:60}},
  xAxis:{{type:'category',data:dates,axisLabel:{{color:'#8e9bbd',rotate:45}}}},
  yAxis:[{{type:'value',axisLabel:{{color:'#8e9bbd',formatter:v=>v+'%'}}}},
         {{type:'value',name:'回撤%',axisLabel:{{color:'#8e9bbd'}},splitLine:{{show:false}}}}],
  series:lsSeries
}});
echarts.init(document.getElementById('chartGroup')).setOption({{
  tooltip:{{trigger:'axis'}}, grid:{{left:60,right:30,top:30,bottom:40}},
  xAxis:{{type:'category',data:GS.map(d=>'G'+d.group),axisLabel:{{color:'#8e9bbd'}}}},
  yAxis:{{type:'value',axisLabel:{{color:'#8e9bbd',formatter:v=>(v*100).toFixed(1)+'%'}}}},
  series:[{{type:'bar',data:GS.map(d=>d.avgReturn),itemStyle:{{color:p=>p.value>=0?'#ff4d4f':'#21c08b'}}}}]
}});
const BE=GS.map(g=>((g.equity||[]).slice(1)).map(v=>Number((v*100).toFixed(2))));
echarts.init(document.getElementById('chartGroupEq')).setOption({{
  tooltip:{{trigger:'axis'}}, legend:{{textStyle:{{color:'#8e9bbd'}},top:0}},
  grid:{{left:60,right:40,top:40,bottom:60}},
  xAxis:{{type:'category',data:BD,axisLabel:{{color:'#8e9bbd',rotate:45}}}},
  yAxis:{{type:'value',axisLabel:{{color:'#8e9bbd',formatter:v=>v+'%'}}}},
  series:GS.map((g,i)=>({{name:'G'+g.group,type:'line',data:BE[i],smooth:true,showSymbol:false}}))
}});
echarts.init(document.getElementById('chartIC')).setOption({{
  tooltip:{{trigger:'axis'}}, legend:{{textStyle:{{color:'#8e9bbd'}},top:0}},
  grid:{{left:60,right:30,top:40,bottom:60}},
  xAxis:{{type:'category',data:IC.map(p=>p.date),axisLabel:{{color:'#8e9bbd',rotate:45}}}},
  yAxis:{{type:'value',axisLabel:{{color:'#8e9bbd'}}}},
  series:[
    {{name:'IC',type:'line',data:IC.map(p=>p.ic),showSymbol:false,itemStyle:{{color:'#4f8cff'}}}},
    {{name:'RankIC',type:'line',data:IC.map(p=>p.rankIc),showSymbol:false,itemStyle:{{color:'#6c5ce7'}}}}
  ]
}});
const FI = {fi_json};
if (FI.length > 0) {{
  document.getElementById('fiCard').style.display = 'block';
  const topN = FI.slice(0, 20).reverse();
  const fiLabel = topN.map(p => p.label || p.feature);
  const fiVal = topN.map(p => Number(p.importance.toFixed ? p.importance.toFixed(4) : p.importance));
  echarts.init(document.getElementById('chartFI')).setOption({{
    tooltip:{{trigger:'axis',axisPointer:{{type:'shadow'}}}},
    grid:{{left:140,right:40,top:10,bottom:30}},
    xAxis:{{type:'value',axisLabel:{{color:'#8e9bbd'}}}},
    yAxis:{{type:'category',data:fiLabel,axisLabel:{{color:'#cfe0ff',fontSize:12}}}},
    series:[{{name:'重要性',type:'bar',data:fiVal,itemStyle:{{color:'#4f8cff'}},label:{{show:true,position:'right',color:'#8e9bbd',fontSize:11}}}}]
  }});
}}
const PL = {_json.dumps(payload.get("positionLedger") or [], ensure_ascii=False)};
if (PL.length > 0) {{
  document.getElementById('chartPosCard').style.display = 'block';
  const posDates = PL.map(p => p.date);
  // 优先用后端直算的持仓数/换手，旧数据回退到前端计算
  const posLongCount = PL.map(p => p.longCount !== undefined ? p.longCount : (p.long ? p.long.length : 0));
  const posShortCount = PL.map(p => p.shortCount !== undefined ? p.shortCount : (p.short ? p.short.length : 0));
  const posTurnover = PL.map((p, i) => {{
    if (p.turnover !== undefined) return p.turnover;
    if (i === 0) return posLongCount[0] + posShortCount[0];
    const prev = new Set([...(PL[i-1].long||[]), ...(PL[i-1].short||[])]);
    const curr = new Set([...(p.long||[]), ...(p.short||[])]);
    let same = 0; prev.forEach(c => {{ if (curr.has(c)) same++; }});
    return (prev.size + curr.size - 2*same);
  }});
  echarts.init(document.getElementById('chartPos')).setOption({{
    tooltip:{{trigger:'axis'}}, legend:{{textStyle:{{color:'#8e9bbd'}},top:0}},
    grid:{{left:60,right:30,top:40,bottom:60}},
    xAxis:{{type:'category',data:posDates,axisLabel:{{color:'#8e9bbd',rotate:45}}}},
    yAxis:[{{type:'value',name:'持仓数',axisLabel:{{color:'#8e9bbd'}}}},{{type:'value',name:'换手数',axisLabel:{{color:'#8e9bbd'}}}}],
    series:[
      {{name:'多头持仓',type:'line',data:posLongCount,showSymbol:false,itemStyle:{{color:'#ff4d4f'}}}},
      {{name:'空头持仓',type:'line',data:posShortCount,showSymbol:false,itemStyle:{{color:'#21c08b'}}}},
      {{name:'换手股数',type:'bar',yAxisIndex:1,data:posTurnover,itemStyle:{{color:'rgba(79,140,255,.35)'}}}}
    ]
  }});
  // 调仓信号明细表（最近 10 期，最新在前）
  const detail = PL.slice(-10).reverse();
  const esc = s => String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');
  const fmtCodes = arr => {{
    const a = arr || [];
    if (!a.length) return '<span style="color:#4a5a78">-</span>';
    const head = a.slice(0,8).map(esc).join(', ');
    return head + (a.length > 8 ? ' <span style="color:#8e9bbd">等' + a.length + '只</span>' : '');
  }};
  const rowsHtml = detail.map(p => {{
    const hasDetail = p.longAdded !== undefined || p.longRemoved !== undefined;
    const la = hasDetail ? fmtCodes(p.longAdded) : '-';
    const lr = hasDetail ? fmtCodes(p.longRemoved) : '-';
    const sa = hasDetail ? fmtCodes(p.shortAdded) : '-';
    const sr = hasDetail ? fmtCodes(p.shortRemoved) : '-';
    return '<tr><td>' + esc(p.date) + '</td><td>' + la + '</td><td>' + lr + '</td><td>' + sa + '</td><td>' + sr + '</td></tr>';
  }}).join('');
  if (rowsHtml) {{
    document.getElementById('rebalanceCard').style.display = 'block';
    document.getElementById('rebalanceTable').innerHTML =
      '<table><tr><th>调仓日</th><th>多头买入</th><th>多头卖出</th><th>空头买入</th><th>空头卖出</th></tr>' + rowsHtml + '</table>';
  }}
}}
</script>
</body></html>"""


def _equity_png(payload: dict) -> bytes | None:
    """用 matplotlib 生成「净值 + 多/空单腿 + 分组净值」PNG，供 Excel/PDF 嵌入。

    matplotlib 未安装时返回 None（报告降级为纯文本）。
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    ls = payload.get("longShort") or []
    gs = payload.get("groupSummary") or []
    if not ls and not gs:
        return None
    direction = payload.get("direction") or "long_short"
    fig, axes = plt.subplots(2, 1, figsize=(9, 7), dpi=110)
    if ls:
        ax = axes[0]
        cum = [p.get("cum", 0.0) * 100 for p in ls]
        ax.plot(range(len(cum)), cum, label="净值", color="#4f8cff", linewidth=1.6)
        if direction == "long_short":
            top = [p.get("topCum", 0.0) * 100 for p in ls]
            bot = [p.get("bottomCum", 0.0) * 100 for p in ls]
            ax.plot(range(len(top)), top, label="多头腿", color="#21c08b", linewidth=1.2)
            ax.plot(range(len(bot)), bot, label="空头腿", color="#6c5ce7", linewidth=1.2)
        ax.set_title("累计净值(%)")
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best", fontsize=8)
    if gs:
        ax = axes[1]
        for g in gs:
            eq = g.get("equity") or []
            ax.plot(range(len(eq) - 1), eq[1:], label=f"G{g.get('group')}", linewidth=1.4)
        ax.set_title("分组净值(%)")
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best", fontsize=8)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110)
    plt.close(fig)
    return buf.getvalue()


def render_excel(payload: dict) -> bytes:
    """生成 Excel（openpyxl）：绩效指标 / 分组收益 / 多空净值 / IC 序列 多 sheet。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ValueError("服务器未安装 openpyxl，无法导出 Excel")
    wb = Workbook()
    m = payload.get("metrics") or {}
    cfg = payload.get("config") or {}

    ws = wb.active
    ws.title = "绩效指标"
    rows = [
        ("因子/模型", payload.get("factorLabel")),
        ("板块", cfg.get("board", "-")),
        ("持有期(n)", cfg.get("n", "-")),
        ("历史长度(hist)", cfg.get("hist", "-")),
        ("分组数", cfg.get("groups", "-")),
        ("候选池规模", cfg.get("poolSize", "-")),
        ("回测起始", cfg.get("startDate", "-")),
        ("回测结束", cfg.get("endDate", "-")),
        ("佣金/印花/滑点", f"{cfg.get('commissionRate',0.00025)}/{cfg.get('stampDuty',0.001)}/{cfg.get('slippage',0.001)}"),
        ("", ""),
        ("累计收益", m.get("cumulativeReturn")),
        ("年化收益", m.get("annualizedReturn")),
        ("年化波动", m.get("annualizedVolatility")),
        ("Sharpe", m.get("sharpe")),
        ("Sortino", m.get("sortino")),
        ("最大回撤", m.get("maxDrawdown")),
        ("Calmar", m.get("calmar")),
        ("胜率", m.get("winRate")),
        ("调仓次数", m.get("rebalanceCount")),
        ("成本率", m.get("costRate")),
        ("交易方向", cfg.get("direction", "long_short")),
        ("年化换手", m.get("annualizedTurnover")),
        ("毛Sharpe(除成本)", m.get("grossSharpe")),
        ("净Sharpe(含成本)", m.get("sharpe")),
        ("成本侵蚀Sharpe", m.get("costErosion")),
    ]
    for r in rows:
        ws.append(r)

    if payload.get("groupSummary"):
        ws2 = wb.create_sheet("分组收益")
        ws2.append(["分组", "平均收益", "累计收益", "年化收益", "Sharpe", "最大回撤", "样本数"])
        for g in payload["groupSummary"]:
            ws2.append([g.get("group"), g.get("avgReturn"), g.get("cumReturn"),
                        g.get("annualizedReturn"), g.get("sharpe"), g.get("maxDrawdown"), g.get("sample")])

    if payload.get("longShort"):
        ws3 = wb.create_sheet("多空净值")
        ws3.append(["日期", "收益", "累计收益", "毛收益", "多头腿累计", "空头腿累计"])
        for p in payload["longShort"]:
            ws3.append([p.get("date"), p.get("longShort"), p.get("cum"),
                        p.get("gross"), p.get("topCum"), p.get("bottomCum")])

    if payload.get("icSeries"):
        ws4 = wb.create_sheet("IC序列")
        ws4.append(["日期", "IC", "RankIC", "样本数"])
        for p in payload["icSeries"]:
            ws4.append([p.get("date"), p.get("ic"), p.get("rankIc"), p.get("sample")])

    if payload.get("yearlyReturns"):
        ws5 = wb.create_sheet("年度绩效")
        ws5.append(["年份", "年度收益", "调仓次数"])
        for y in payload["yearlyReturns"]:
            ws5.append([y.get("year"), y.get("return"), y.get("periods")])

    if payload.get("stockContribution"):
        ws6 = wb.create_sheet("个股贡献")
        ws6.append(["代码", "名称", "累计贡献"])
        for sc in payload["stockContribution"]:
            ws6.append([sc.get("code"), sc.get("name"), sc.get("contribution")])

    # B9：图表 PNG 嵌入
    png = _equity_png(payload)
    if png:
        try:
            from openpyxl.drawing.image import Image as XlImage
            ws7 = wb.create_sheet("图表")
            img = XlImage(io.BytesIO(png))
            img.width = 640
            img.height = 480
            ws7.add_image(img, "A1")
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_pdf(payload: dict) -> bytes:
    """生成 PDF（reportlab CID 字体 STSong-Light，中文无需外挂字体文件）。

    reportlab 未安装时返回 None，由调用方降级为打印样式 HTML。
    """
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    except ImportError:
        return None
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    y = h - 50
    m = payload.get("metrics") or {}
    bench = payload.get("benchmark")

    def line():
        nonlocal y
        c.setFillColorRGB(0.18, 0.22, 0.30)
        c.rect(50, y - 2, w - 100, 0.5, stroke=0, fill=1)
        y -= 14

    def kpi_row(label, val):
        nonlocal y
        if y < 60:
            c.showPage(); y = h - 50
        c.setFont("STSong-Light", 10)
        c.setFillColorRGB(0.15, 0.15, 0.25)
        c.drawString(70, y, label)
        c.setFillColorRGB(0.05, 0.2, 0.6)
        c.drawRightString(w - 70, y, _fmt(val))
        y -= 18

    c.setFont("STSong-Light", 18)
    c.setFillColorRGB(0.05, 0.1, 0.25)
    c.drawString(50, y, payload.get("factorLabel") or "回测报告")
    y -= 22
    c.setFont("STSong-Light", 9)
    c.setFillColorRGB(0.4, 0.4, 0.5)
    c.drawString(50, y, f"生成时间 {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    y -= 14
    # 回测区间与策略来源（HTML 与 PDF 均含，补齐报告历史缺口）
    cfg = payload.get("config") or {}
    sdate, edate = cfg.get("startDate") or "最远", cfg.get("endDate") or "最新"
    strategy_kind = "ML模型" if cfg.get("modelId") else "技术因子"
    strategy_name = cfg.get("modelId") or payload.get("factorLabel") or "-"
    c.drawString(50, y, f"回测区间 {sdate} ~ {edate}  |  策略来源 {strategy_kind}  |  {strategy_name}")
    y -= 26
    c.setFont("STSong-Light", 13)
    c.drawString(50, y, "绩效指标")
    y -= 20
    for lab, key in (("累计收益", "cumulativeReturn"), ("年化收益", "annualizedReturn"),
                     ("年化波动", "annualizedVolatility"), ("Sharpe", "sharpe"),
                     ("Sortino", "sortino"), ("最大回撤", "maxDrawdown"),
                     ("Calmar", "calmar"), ("胜率", "winRate"),
                     ("调仓次数", "rebalanceCount"), ("成本率", "costRate"),
                     ("交易方向", "direction"), ("年化换手", "annualizedTurnover"),
                     ("毛Sharpe(除成本)", "grossSharpe"), ("成本侵蚀Sharpe", "costErosion")):
        kpi_row(lab, m.get(key) if key != "direction" else (payload.get("direction") or cfg.get("direction") or "long_short"))
    y -= 6

    if bench:
        if y < 90:
            c.showPage(); y = h - 50
        c.setFont("STSong-Light", 13)
        c.drawString(50, y, "基准对比")
        y -= 20
        for lab, key in (("累计收益", "cumulativeReturn"), ("年化收益", "annualizedReturn"),
                         ("Sharpe", "sharpe")):
            kpi_row(lab, bench.get(key))
        # 策略侧 Alpha/Beta（B1：与 HTML 一致，取 strategyAlpha/strategyBeta）
        for lab, key in (("策略 Alpha", "strategyAlpha"), ("策略 Beta", "strategyBeta")):
            val = bench.get(key)
            if val is None:
                val = bench.get("alpha" if key == "strategyAlpha" else "beta")
            kpi_row(lab, val)
        y -= 6

    if payload.get("groupSummary"):
        if y < 120:
            c.showPage(); y = h - 50
        c.setFont("STSong-Light", 13)
        c.drawString(50, y, "分组收益")
        y -= 20
        for g in payload["groupSummary"]:
            kpi_row(f"第{g.get('group')}组", g.get("avgReturn"))
        y -= 6

    if payload.get("icSeries"):
        ics = [p.get("ic") for p in payload["icSeries"] if p.get("ic") is not None]
        if ics:
            mean_ic = sum(ics) / len(ics)
            pos = sum(1 for v in ics if v > 0) / len(ics)
            c.setFont("STSong-Light", 10)
            c.setFillColorRGB(0.15, 0.15, 0.25)
            c.drawString(70, y, f"IC 统计：平均 IC {mean_ic:.4f} · 胜率 {pos*100:.1f}% · 截面 {len(ics)}")
            y -= 24

    if payload.get("yearlyReturns"):
        if y < 120:
            c.showPage(); y = h - 50
        c.setFont("STSong-Light", 13)
        c.drawString(50, y, "分年度绩效")
        y -= 20
        for yr in payload["yearlyReturns"]:
            kpi_row(f"{yr.get('year')} 年", yr.get("return"))
        y -= 6

    # B5：个股贡献 Top（HTML/Excel 已有，补齐 PDF）
    stock_contrib = payload.get("stockContribution") or []
    if stock_contrib:
        if y < 120:
            c.showPage(); y = h - 50
        c.setFont("STSong-Light", 13)
        c.drawString(50, y, "个股贡献 Top30")
        y -= 20
        for sc in stock_contrib[:30]:
            name = sc.get("name") or ""
            kpi_row(f"{sc.get('code')} {name}".strip(), sc.get("contribution"))
        y -= 6

    if y < 60:
        c.showPage(); y = h - 50
    c.setFont("STSong-Light", 9)
    c.setFillColorRGB(0.5, 0.5, 0.55)
    c.drawString(50, y, "说明：候选池为当前上市标的快照，历史收益可能存在幸存者偏差；"
                        "本报告由量化研究平台自动生成。")
    y -= 30

    # B9：图表 PNG 嵌入（matplotlib 可用时）
    png = _equity_png(payload)
    if png:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(io.BytesIO(png))
            iw, ih = img.getSize()
            scale = min((w - 100) / iw, 260 / ih)
            dw, dh = iw * scale, ih * scale
            if y < dh + 40:
                c.showPage(); y = h - 50
            c.drawImage(img, 50, y - dh, width=dw, height=dh)
        except Exception:
            pass

    c.save()
    return buf.getvalue()


def store_backtest_report(result: dict, config: dict | None = None,
                          user_id: int = 0) -> dict | None:
    """回测完成自动存档：生成 HTML 报告文件并登记 backtest_runs 记录。

    返回 run 记录 dict；失败（无 longShort 数据等）返回 None，不阻塞回测主流程。
    """
    try:
        payload = payload_from_result(result)
        if config is not None:
            result["config"] = config
            payload["config"] = config
        if not payload["longShort"] and not payload["icSeries"]:
            return None
        html = render_html(payload)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"report_{ts}_{uuid.uuid4().hex[:6]}.html"
        path = os.path.join(REPORT_DIR, fname)
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        run = db.create_backtest_run(
            None, config or result.get("config") or {},
            result.get("metrics") or {}, report_path=fname,
            user_id=user_id, result=result)
        return run
    except Exception as e:
        logger.warning("回测报告自动存档失败: %s", e)
        return None


def regenerate_report(run: dict, fmt: str = "html"):
    """从历史 run 的 result 重新生成指定格式报告，返回 (bytes, media_type)。"""
    result = run.get("result") or {}
    if not result:
        return None, None
    payload = payload_from_result(result)
    if fmt == "excel":
        return render_excel(payload), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if fmt == "pdf":
        data = render_pdf(payload)
        if data:
            return data, "application/pdf"
        return None, None
    return render_html(payload), "text/html"

